#!/usr/bin/env python3
"""
生成贴图贡献追踪表（多工作表 xlsx）—— 适配 CCB 新物品格式

直接扫描 data/json 采集所有需要贴图的实体，按类别分到不同工作表，
标注每个贴图已有/缺失，并附协作空列（提交/贡献者/完成状态），供团队认领绘制。

为什么不用官方 table.py --tileset / generate_overlay_ids.py：
  CCB（本分支）已迁移到新物品格式 "type": "ITEM" + subtypes:[ARMOR/GUN/...]，
  而官方工具按旧的 "type": "ARMOR"/"GUN" 匹配，会漏掉 CCB 几乎全部物品。
  本脚本按新格式直接扫描，保证护甲/武器/工具等正确分类。

工作表划分：
  汇总            —— 各表统计概览
  主贴图-物品     —— 所有物品本体贴图
  主贴图-怪物/地形/家具/变异生化/载具/陷阱场效 —— 各实体本体
  穿戴-中性       —— 所有护甲应有的 overlay_worn_
  穿戴-男 / 穿戴-女 —— 真实贴图中已存在的男女穿戴版
  手持            —— 所有可手持物品应有的 overlay_wielded_
  变异生化叠加    —— overlay_mutation_（含男女已有版标注）
  尸体            —— corpse_
  overmap         —— overmap_terrain / overmap_special
  地图事件        —— map_extra (mx_)

依赖 openpyxl。tileset_ids.csv 来自 list_tileset_ids.py（compose 后的成品目录）。

用法：
  python3 tools/gfx_tools/list_tileset_ids.py <composed_dir> > tileset_ids.csv
  python3 tools/gfx_tools/tileset_workbook.py tileset_ids.csv \\
    --data-dir data/json --output 贴图贡献追踪表.xlsx
"""
import argparse
import csv
import glob
import json
import os

COLUMNS = ['类型', 'ID', '名称', '前缀', '完整贴图ID',
           '当前贴图', '完成状态', '提交', '贡献者', '备注']

SPRITE_HAVE = '已有'
SPRITE_MISSING = '缺失'
STATUS_MERGED = '已合并'
STATUS_TODO = '待领取'

OVERMAP_TYPES = ('overmap_terrain', 'overmap_special')
MAP_EXTRA_TYPE = 'map_extra'

# CCB 新格式：物品统一 type:ITEM，细分类靠 subtypes
ITEM_WORN_SUBTYPES = ('ARMOR', 'TOOL_ARMOR', 'PET_ARMOR')
ITEM_WIELDED_SUBTYPES = ('GUN', 'TOOL', 'GUNMOD', 'TOOLMOD', 'AMMO',
                         'MAGAZINE', 'BATTERY', 'ENGINE', 'WHEEL', 'BOOK',
                         'COMESTIBLE', 'BIONIC_ITEM')

SIMPLE_TYPE_MAP = {
    'MONSTER': 'monsters', 'terrain': 'terrain', 'furniture': 'furniture',
    'mutation': 'mutation', 'bionic': 'mutation', 'vehicle_part': 'vehicle_part',
    'field_type': 'field', 'trap': 'trap', 'gate': 'gate', 'SPELL': 'spell',
}


def get_name(o):
    name = o.get('name')
    if isinstance(name, dict):
        return name.get('str') or name.get('str_sp') or ''
    return str(name or '')


def load_csv_ids(path):
    out = []
    with open(path, newline='', encoding='utf-8') as f:
        for row in csv.reader(f):
            if row and row[0].strip():
                out.append(row[0].strip())
    return out


def scan_game_data(data_dir):
    """按 CCB 新格式直接扫 data/json，返回各类别的 [(id,name)]。"""
    buckets = {k: [] for k in (
        'items_worn', 'items_wielded', 'items_other', 'monsters', 'terrain',
        'furniture', 'mutation', 'vehicle_part', 'field', 'trap', 'gate',
        'spell', 'overmap', 'map_extra')}
    seen = {k: set() for k in buckets}

    def add(bucket, oid, name):
        if oid and oid not in seen[bucket]:
            seen[bucket].add(oid)
            buckets[bucket].append((oid, name))

    for fp in glob.glob(os.path.join(data_dir, '**', '*.json'), recursive=True):
        try:
            with open(fp, encoding='utf-8') as f:
                data = json.load(f)
        except Exception:
            continue
        if isinstance(data, dict):
            data = [data]
        if not isinstance(data, list):
            continue
        for o in data:
            if not isinstance(o, dict) or o.get('abstract'):
                continue
            t = o.get('type')
            oid = o.get('id')
            if isinstance(oid, list):
                oid = oid[0] if oid else None
            name = get_name(o)

            if t == 'ITEM':
                if not oid:
                    continue
                subs = o.get('subtypes') or []
                if any(s in subs for s in ITEM_WORN_SUBTYPES):
                    add('items_worn', oid, name)
                elif any(s in subs for s in ITEM_WIELDED_SUBTYPES):
                    add('items_wielded', oid, name)
                else:
                    add('items_other', oid, name)
            elif t in SIMPLE_TYPE_MAP:
                add(SIMPLE_TYPE_MAP[t], oid, name)
            elif t in OVERMAP_TYPES:
                om_id = oid or o.get('om_terrain')
                if isinstance(om_id, list):
                    om_id = om_id[0] if om_id else None
                if om_id and om_id not in seen['overmap']:
                    seen['overmap'].add(om_id)
                    buckets['overmap'].append((str(om_id), name, t))
            elif t == MAP_EXTRA_TYPE:
                add('map_extra', oid, name)

    return buckets


def make_row(type_label, gid, name, prefix, full_id, has_sprite):
    return {
        '类型': type_label, 'ID': gid, '名称': name,
        '前缀': prefix, '完整贴图ID': full_id,
        '当前贴图': SPRITE_HAVE if has_sprite else SPRITE_MISSING,
        '完成状态': STATUS_MERGED if has_sprite else STATUS_TODO,
        '提交': '', '贡献者': '', '备注': '',
    }


def sort_rows(rows):
    return sorted(rows, key=lambda r: (r['当前贴图'] != SPRITE_MISSING,
                                       r['类型'], r['ID']))


def build_sheets(buckets, tileset_ids):
    ts = set(tileset_ids)
    name_lookup = {}
    for b in ('items_worn', 'items_wielded', 'items_other', 'monsters',
              'terrain', 'furniture', 'mutation', 'vehicle_part'):
        for gid, name in buckets[b]:
            name_lookup[gid] = name

    sheets = {}

    # ---- 主贴图：各实体本体（id 直接作为贴图 id） ----
    def body_sheet(label, bucket):
        return sort_rows([make_row(label, gid, name, '', gid, gid in ts)
                          for gid, name in buckets[bucket]])

    item_body = []
    for bucket, lab in (('items_worn', '护甲'), ('items_wielded', '武器工具'),
                        ('items_other', '杂项物品')):
        for gid, name in buckets[bucket]:
            item_body.append(make_row(lab, gid, name, '', gid, gid in ts))
    sheets['主贴图-物品'] = sort_rows(item_body)
    sheets['主贴图-怪物'] = body_sheet('MONSTER', 'monsters')
    sheets['主贴图-地形'] = body_sheet('terrain', 'terrain')
    sheets['主贴图-家具'] = body_sheet('furniture', 'furniture')
    # 注：mutation/bionic 无本体贴图类别，只通过 overlay_mutation_ 显示，
    # 故不单列「主贴图-变异生化」，统一放到下方「变异生化叠加」表。
    sheets['主贴图-载具部件'] = body_sheet('vehicle_part', 'vehicle_part')
    other_body = []
    for bucket, lab in (('field', 'field_type'), ('trap', 'trap'),
                        ('gate', 'gate'), ('spell', 'SPELL')):
        for gid, name in buckets[bucket]:
            other_body.append(make_row(lab, gid, name, '', gid, gid in ts))
    sheets['主贴图-陷阱场效'] = sort_rows(other_body)

    # ---- 穿戴 overlay：所有护甲都应有 overlay_worn_ ----
    worn = []
    for gid, name in buckets['items_worn']:
        full = 'overlay_worn_' + gid
        worn.append(make_row('穿戴', gid, name, 'overlay_worn_', full, full in ts))
    sheets['穿戴-中性'] = sort_rows(worn)

    # 男/女穿戴：真实贴图里已存在的版本（已完成的精细化工作）
    def gender_sheet(gender_prefix, label):
        rows = []
        for full in ts:
            if full.startswith(gender_prefix):
                base = full[len(gender_prefix):]
                rows.append(make_row(label, base, name_lookup.get(base, ''),
                                     gender_prefix, full, True))
        return sort_rows(rows)
    sheets['穿戴-男'] = gender_sheet('overlay_male_worn_', '穿戴-男')
    sheets['穿戴-女'] = gender_sheet('overlay_female_worn_', '穿戴-女')

    # ---- 手持 overlay：可手持物品应有 overlay_wielded_ ----
    wielded = []
    for gid, name in buckets['items_wielded']:
        full = 'overlay_wielded_' + gid
        wielded.append(make_row('手持', gid, name, 'overlay_wielded_', full, full in ts))
    sheets['手持'] = sort_rows(wielded)

    # ---- 变异/生化 overlay ----
    mut = []
    for gid, name in buckets['mutation']:
        full = 'overlay_mutation_' + gid
        mut.append(make_row('变异生化叠加', gid, name, 'overlay_mutation_', full, full in ts))
    sheets['变异生化叠加'] = sort_rows(mut)

    # ---- 尸体 corpse_ ----
    corpse = []
    for gid, name in buckets['monsters']:
        full = 'corpse_' + gid
        corpse.append(make_row('尸体', gid, name, 'corpse_', full, full in ts))
    sheets['尸体'] = sort_rows(corpse)

    # ---- overmap ----
    om = []
    for gid, name, t in buckets['overmap']:
        has = gid in ts or any(x == gid or x.startswith(gid + '_') for x in ts)
        om.append(make_row(t, gid, name, '', gid, has))
    sheets['overmap'] = sort_rows(om)

    # ---- 地图事件 map_extra ----
    mx = []
    for gid, name in buckets['map_extra']:
        # map_extra 的 id 本身已带 mx_ 前缀（如 mx_helicopter），不重复添加
        full = gid if gid.startswith('mx_') else 'mx_' + gid
        mx.append(make_row('map_extra', gid, name, 'mx_', full, full in ts))
    sheets['地图事件'] = sort_rows(mx)

    return sheets


SHEET_ORDER = [
    '主贴图-物品', '主贴图-怪物', '主贴图-地形', '主贴图-家具',
    '主贴图-载具部件', '主贴图-陷阱场效',
    '穿戴-中性', '穿戴-男', '穿戴-女', '手持',
    '变异生化叠加', '尸体', 'overmap', '地图事件',
]


def write_workbook(sheets, output_path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill('solid', fgColor='4472C4')
    miss_fill = PatternFill('solid', fgColor='FFE699')
    center = Alignment(horizontal='center', vertical='center')

    summary = wb.create_sheet('汇总')
    summary.append(['工作表', '总数', '已有', '缺失', '覆盖率'])
    for c in summary[1]:
        c.font, c.fill, c.alignment = header_font, header_fill, center

    for name in SHEET_ORDER:
        rows = sheets.get(name, [])
        ws = wb.create_sheet(name)
        ws.append(COLUMNS)
        for c in ws[1]:
            c.font, c.fill, c.alignment = header_font, header_fill, center
        ws.freeze_panes = 'A2'
        for r in rows:
            ws.append([r[k] for k in COLUMNS])
            if r['当前贴图'] == SPRITE_MISSING:
                for c in ws[ws.max_row]:
                    c.fill = miss_fill
        for col, w in {'A': 16, 'B': 32, 'C': 24, 'D': 18, 'E': 36,
                       'F': 9, 'G': 9, 'H': 24, 'I': 12, 'J': 16}.items():
            ws.column_dimensions[col].width = w
        total = len(rows)
        have = sum(1 for r in rows if r['当前贴图'] == SPRITE_HAVE)
        cov = f'{have/total*100:.1f}%' if total else '—'
        summary.append([name, total, have, total - have, cov])

    for col, w in {'A': 18, 'B': 10, 'C': 10, 'D': 10, 'E': 12}.items():
        summary.column_dimensions[col].width = w

    wb.save(output_path)


if __name__ == '__main__':
    ap = argparse.ArgumentParser(
        description=__doc__,
        formatter_class=argparse.RawDescriptionHelpFormatter)
    ap.add_argument('tileset_ids_csv', help='list_tileset_ids.py 的输出')
    ap.add_argument('--data-dir', default='data/json')
    ap.add_argument('--output', default='贴图贡献追踪表.xlsx')
    args = ap.parse_args()

    tileset_ids = load_csv_ids(args.tileset_ids_csv)
    buckets = scan_game_data(args.data_dir)
    sheets = build_sheets(buckets, tileset_ids)
    write_workbook(sheets, args.output)

    print(f'✓ 已生成 {args.output}')
    grand = 0
    for name in SHEET_ORDER:
        rows = sheets.get(name, [])
        grand += len(rows)
        miss = sum(1 for r in rows if r['当前贴图'] == SPRITE_MISSING)
        print(f'    {name}: {len(rows)} 条（缺失 {miss}）')
    print(f'  合计 {grand} 条')
